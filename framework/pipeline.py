"""
Evaluation Pipeline Orchestration.

Provides high-level API for running evaluations end-to-end.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import warnings

from .config_loader import load_mode_config
from .scripts.gt_loader import GTLoader
from .validators.pre_eval import validate_pre_evaluation
from .validators.pre_aggregate import validate_pre_aggregation
from .validators.pre_workbook import validate_pre_workbook

logger = logging.getLogger(__name__)


class EvaluationPipeline:
    """
    Orchestrates evaluation pipeline from GT loading through workbook generation.

    Usage:
        pipeline = EvaluationPipeline(mode="freeform", mode_dir=Path("freeform"))
        pipeline.validate_prerequisites("pre_eval", env="hotfix")
        result = pipeline.score_evaluation("consulting", "pathfinder", canonical_json_path)
        summary = pipeline.run_full_pipeline(env="hotfix")
    """

    def __init__(
        self,
        mode: str,
        mode_dir: Optional[Path] = None,
        config_path: Optional[Path] = None
    ):
        """
        Initialize pipeline for a specific evaluation mode.

        Args:
            mode: Mode name (freeform, rules, guidelines, etc.)
            mode_dir: Path to mode directory. If None, infers from config.
            config_path: Path to mode config JSON. If None, uses default location.
        """
        self.mode = mode

        # Load and validate configuration
        if config_path is None:
            config_path = Path("framework/config") / f"{mode}.json"

        self.config = load_mode_config(config_path, validate=True)

        # Set mode directory
        if mode_dir is None:
            base_dir = self.config.get("paths", {}).get("base_dir", mode)
            mode_dir = Path(base_dir)

        self.mode_dir = Path(mode_dir)

        if not self.mode_dir.exists():
            raise FileNotFoundError(f"Mode directory not found: {self.mode_dir}")

        # Initialize GT loader
        self.gt_loader = GTLoader(self.mode_dir, self.config)

    def load_ground_truth(
        self,
        contract: str,
        contract_type: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Load ground truth for a contract using mode-specific structure.

        Args:
            contract: Contract identifier
            contract_type: Contract type (for modes organized by type)

        Returns:
            Dictionary containing GT data with resolved fields
        """
        result = self.gt_loader.load(contract, contract_type=contract_type)
        return {
            "data": result.data,
            "gt_type": result.gt_type,
            "source_files": [str(f) for f in result.source_files]
        }

    def validate_prerequisites(
        self,
        stage: str,
        env: Optional[str] = None,
        run_dirs: Optional[List[Path]] = None
    ) -> None:
        """
        Run validation gate for a specific pipeline stage.

        Args:
            stage: Stage to validate (pre_eval, pre_aggregate, pre_workbook)
            env: Environment name (required for pre_eval)
            run_dirs: Run directories (required for pre_aggregate, pre_workbook)

        Raises:
            ValidationError: If validation fails with ERROR-severity issues
        """
        if stage == "pre_eval":
            if env is None:
                raise ValueError("env parameter required for pre_eval validation")

            result = validate_pre_evaluation(self.mode_dir, env)
            result.abort_if_errors("pre-evaluation")

            if result.warnings:
                for warning in result.warnings:
                    warnings.warn(f"{warning.message} (at {warning.location})", UserWarning)

        elif stage == "pre_aggregate":
            if run_dirs is None:
                raise ValueError("run_dirs parameter required for pre_aggregate validation")

            result = validate_pre_aggregation(run_dirs)
            result.abort_if_errors("pre-aggregation")

            if result.warnings:
                for warning in result.warnings:
                    warnings.warn(f"{warning.message} (at {warning.location})", UserWarning)

        elif stage == "pre_workbook":
            if env is None:
                raise ValueError("env parameter required for pre_workbook validation")

            # pre_workbook validates mode_dir for aggregated results
            result = validate_pre_workbook(self.mode_dir, env)
            result.abort_if_errors("pre-workbook")

            if result.warnings:
                for warning in result.warnings:
                    warnings.warn(f"{warning.message} (at {warning.location})", UserWarning)

        else:
            raise ValueError(f"Unknown validation stage: {stage}")

    def validate_runs(self, run_dirs: List[Path]) -> None:
        """
        Validate multiple evaluation runs for aggregation.

        Args:
            run_dirs: List of run directory paths

        Raises:
            ValidationError: If coverage is incomplete or inconsistent
        """
        self.validate_prerequisites("pre_aggregate", run_dirs=run_dirs)

    def score_evaluation(
        self,
        contract: str,
        model: str,
        canonical_json_path: Path,
        contract_type: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Score a single evaluation using mode-specific scoring logic.

        Current implementation: Loads pre-scored evaluation JSON.
        Future: Implement raw model output → scored evaluation transformation.

        Args:
            contract: Contract identifier
            model: Model identifier
            canonical_json_path: Path to scored evaluation JSON
                (currently expects already-scored evaluation with gt_evaluations array)
            contract_type: Contract type (for rules/guidelines modes)

        Returns:
            Dictionary containing scored evaluation with summary

        Raises:
            FileNotFoundError: If canonical JSON file doesn't exist
            json.JSONDecodeError: If file contains invalid JSON
            KeyError: If required fields missing from scored evaluation
        """
        # Load GT for validation
        gt_result = self.load_ground_truth(contract, contract_type=contract_type)
        gt_issues = gt_result["data"].get("ground_truth", [])

        # Load scored evaluation (canonical JSON)
        try:
            with open(canonical_json_path) as f:
                scored_eval = json.load(f)
        except FileNotFoundError:
            raise FileNotFoundError(
                f"Scored evaluation not found: {canonical_json_path}\n"
                f"Expected structure: {{contract}}/{{model}}.json"
            )
        except json.JSONDecodeError as e:
            raise json.JSONDecodeError(
                f"Invalid JSON in scored evaluation: {canonical_json_path}",
                e.doc, e.pos
            )

        # Validate structure
        if "gt_evaluations" not in scored_eval:
            raise KeyError(
                f"Scored evaluation missing 'gt_evaluations' array: {canonical_json_path}\n"
                f"Found keys: {list(scored_eval.keys())}"
            )

        # Verify GT coverage
        eval_gt_ids = {e.get("gt_id") for e in scored_eval["gt_evaluations"]}
        expected_gt_ids = {gt.get("gt_id") for gt in gt_issues}

        missing = expected_gt_ids - eval_gt_ids
        if missing:
            logger.warning(
                f"Scored evaluation missing GT IDs: {missing}\n"
                f"Contract: {contract}, Model: {model}"
            )

        # Add summary if not present
        if "summary" not in scored_eval:
            scored_eval["summary"] = self._calculate_summary(
                scored_eval["gt_evaluations"],
                gt_issues
            )

        return scored_eval

    def _calculate_summary(
        self,
        gt_evaluations: List[Dict[str, Any]],
        gt_issues: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Calculate summary statistics for scored evaluation."""
        from framework.scoring import calculate_detection_points

        tier_config = self.config.get("detection_points", {})

        total_detection_points = 0.0
        total_quality_points = 0.0
        detection_counts = {"Y": 0, "P": 0, "N": 0, "NMI": 0}
        t1_gate_pass = True

        for eval_item in gt_evaluations:
            # Detection points
            detection = eval_item.get("detection", "NMI")
            tier = eval_item.get("tier", "T3")
            detection_points = calculate_detection_points(detection, tier, tier_config)
            total_detection_points += detection_points

            # Quality points
            quality_dims = self.config.get("quality_scores", {}).get("dimensions", [])
            for dim in quality_dims:
                total_quality_points += eval_item.get(dim, 0)

            # Detection counts
            if detection in detection_counts:
                detection_counts[detection] += 1

            # T1 gate check
            if tier == "T1" and detection not in ("Y", "P"):
                t1_gate_pass = False

        # Calculate max possible points
        max_detection_points = sum(
            tier_config.get(gt.get("tier", "T3"), {}).get("Y", 0)
            for gt in gt_issues
        )

        return {
            "total_detection_points": total_detection_points,
            "max_detection_points": max_detection_points,
            "total_quality_points": total_quality_points,
            "total_points": total_detection_points + total_quality_points,
            "detection_counts": detection_counts,
            "t1_gate_pass": t1_gate_pass,
            "weighted_recall": total_detection_points / max_detection_points if max_detection_points > 0 else 0.0
        }

    def aggregate_results(
        self,
        run_dirs: List[Path],
        output_dir: Path
    ) -> Dict[str, Any]:
        """
        Aggregate evaluations from multiple runs.

        Args:
            run_dirs: List of run directories to aggregate
            output_dir: Directory to write aggregated results

        Returns:
            Dictionary with aggregation summary
        """
        # Validate runs first
        logger.info(f"Validating {len(run_dirs)} runs before aggregation")
        self.validate_runs(run_dirs)

        # Collect all evaluations by (contract, model) key
        aggregated = {}
        run_count = 0

        for run_dir in run_dirs:
            eval_dir = run_dir / "evaluations"
            if not eval_dir.exists():
                logger.warning(f"Evaluation directory not found: {eval_dir}, skipping")
                continue

            run_count += 1
            logger.info(f"Processing run {run_count}: {run_dir}")

            for contract_dir in eval_dir.iterdir():
                if not contract_dir.is_dir():
                    continue

                contract = contract_dir.name

                for model_file in contract_dir.glob("*.json"):
                    model = model_file.stem
                    key = (contract, model)

                    try:
                        with open(model_file) as f:
                            evaluation = json.load(f)
                    except (json.JSONDecodeError, OSError) as e:
                        logger.error(f"Failed to load {model_file}: {e}")
                        continue

                    if key not in aggregated:
                        aggregated[key] = {
                            "runs": [],
                            "contracts_models": key
                        }

                    aggregated[key]["runs"].append({
                        "run_dir": str(run_dir),
                        "evaluation": evaluation
                    })

        if not aggregated:
            logger.warning("No evaluations found in any run directory")
            return {
                "files_written": 0,
                "contracts": [],
                "models": [],
                "message": "No evaluations to aggregate"
            }

        # Write aggregated results
        output_dir.mkdir(parents=True, exist_ok=True)
        contracts_processed = set()
        models_processed = set()

        for (contract, model), data in aggregated.items():
            contract_dir = output_dir / contract
            contract_dir.mkdir(exist_ok=True)

            output_file = contract_dir / f"{model}.json"

            # Take the most recent evaluation (last run) as primary
            primary_eval = data["runs"][-1]["evaluation"]

            # Add aggregation metadata
            aggregated_result = {
                **primary_eval,
                "aggregation_meta": {
                    "num_runs": len(data["runs"]),
                    "run_dirs": [r["run_dir"] for r in data["runs"]],
                    "aggregated_at": datetime.now().isoformat()
                }
            }

            with open(output_file, "w") as f:
                json.dump(aggregated_result, f, indent=2)

            contracts_processed.add(contract)
            models_processed.add(model)

        logger.info(
            f"Aggregation complete: {len(aggregated)} files written, "
            f"{len(contracts_processed)} contracts, {len(models_processed)} models"
        )

        return {
            "files_written": len(aggregated),
            "contracts": sorted(contracts_processed),
            "models": sorted(models_processed),
            "output_dir": str(output_dir)
        }

    def generate_workbook(
        self,
        aggregated_dir: Path,
        output_path: Path,
        env: str
    ) -> Path:
        """
        Generate Excel workbook from aggregated results.

        Args:
            aggregated_dir: Directory containing aggregated JSON results
            output_path: Path for output workbook
            env: Environment name

        Returns:
            Path to generated workbook
        """
        # Validate prerequisites
        logger.info(f"Validating aggregated results in {aggregated_dir}")
        self.validate_prerequisites("pre_workbook", env=env)

        # Import workbook generation dependencies
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl required for workbook generation. "
                "Install with: pip install openpyxl"
            )

        # Discover contracts and models from aggregated directory
        contracts = []
        models = set()

        for item in sorted(aggregated_dir.iterdir()):
            if item.is_dir():
                contracts.append(item.name)
                for json_file in item.glob("*.json"):
                    models.add(json_file.stem)

        models = sorted(models)

        logger.info(
            f"Found {len(contracts)} contracts and {len(models)} models "
            f"in {aggregated_dir}"
        )

        if not contracts or not models:
            raise ValueError(
                f"No evaluations found in {aggregated_dir}. "
                "Expected structure: contract_name/model_name.json"
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Styles
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Write header
        ws["A1"] = f"{self.config['display_name']} Evaluation - {env}"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Summary"
        ws["A3"].font = Font(bold=True, size=12)

        # Write contracts and models summary
        row = 5
        ws[f"A{row}"] = "Contracts:"
        ws[f"B{row}"] = ", ".join(contracts)
        row += 1

        ws[f"A{row}"] = "Models:"
        ws[f"B{row}"] = ", ".join(models)
        row += 2

        ws[f"A{row}"] = "Contract"
        ws[f"B{row}"] = "Model"
        ws[f"C{row}"] = "Detection Points"
        ws[f"D{row}"] = "T1 Gate"

        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{row}"]
            cell.fill = header_fill
            cell.font = header_font

        row += 1

        # Load and summarize each evaluation
        for contract in contracts:
            for model in models:
                eval_path = aggregated_dir / contract / f"{model}.json"
                if not eval_path.exists():
                    continue

                with open(eval_path) as f:
                    evaluation = json.load(f)

                summary = evaluation.get("summary", {})

                ws[f"A{row}"] = contract
                ws[f"B{row}"] = model
                ws[f"C{row}"] = summary.get("total_detection_points", 0)
                ws[f"D{row}"] = "PASS" if summary.get("t1_gate_pass", False) else "FAIL"

                row += 1

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Workbook generated: {output_path}")

        return output_path

    def run_full_pipeline(
        self,
        env: str,
        run_dirs: Optional[List[Path]] = None,
        output_dir: Optional[Path] = None
    ) -> Dict[str, Any]:
        """
        Run complete pipeline: validate → aggregate → workbook.

        Args:
            env: Environment name
            run_dirs: Run directories to aggregate (auto-discovers if None)
            output_dir: Output directory (defaults to mode_dir/results)

        Returns:
            Dictionary with execution summary
        """
        # Auto-discover runs if not provided
        if run_dirs is None:
            env_dir = self.mode_dir / "environments" / env
            if env_dir.exists():
                run_dirs = [
                    d for d in env_dir.iterdir()
                    if d.is_dir() and (d / "evaluations").exists()
                ]
            else:
                # Legacy structure
                run_dirs = [
                    d for d in self.mode_dir.iterdir()
                    if d.is_dir() and d.name.startswith("run") and (d / "evaluations").exists()
                ]

        if not run_dirs:
            raise ValueError(f"No evaluation runs found for environment: {env}")

        # Set output directory
        if output_dir is None:
            output_dir = self.mode_dir / "results"

        output_dir = Path(output_dir)

        # Step 1: Validate
        print(f"Validating {len(run_dirs)} runs...")
        self.validate_prerequisites("pre_eval", env=env)
        self.validate_runs(run_dirs)

        # Step 2: Aggregate
        print(f"Aggregating results to {output_dir}...")
        agg_summary = self.aggregate_results(run_dirs, output_dir)

        # Step 3: Generate workbook
        workbook_path = output_dir / f"{self.mode}_{env}.xlsx"
        print(f"Generating workbook: {workbook_path}...")
        self.generate_workbook(output_dir, workbook_path, env)

        return {
            "mode": self.mode,
            "env": env,
            "runs_processed": len(run_dirs),
            "output_dir": str(output_dir),
            "workbook": str(workbook_path),
            "status": "completed"
        }
