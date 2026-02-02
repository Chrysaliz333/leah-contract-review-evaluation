#!/usr/bin/env python3
"""
Generate evaluation workbooks from aggregated data.

Creates:
- Per-contract workbooks: {CONTRACT}_Evaluations.xlsx
- Master workbook: MASTER_EVALUATION_WORKBOOK.xlsx (optional)

Usage:
    python generate_workbooks.py --base-path /path/to/aggregated
    python generate_workbooks.py --base-path /path/to/aggregated --contract consulting
    python generate_workbooks.py --base-path /path/to/aggregated --master-only
"""

import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip3 install openpyxl")
    exit(1)

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
T1_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
T2_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
T3_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

TIER_FILLS = {'T1': T1_FILL, 'T2': T2_FILL, 'T3': T3_FILL}
DETECTION_FILLS = {'Y': PASS_FILL, '△': PARTIAL_FILL, 'P': PARTIAL_FILL, 'N': FAIL_FILL, 'NMI': FAIL_FILL}


def discover_structure(base_path: Path):
    """Discover contracts and models from directory structure."""
    contracts = []
    models = set()
    
    for item in sorted(base_path.iterdir()):
        if item.is_dir():
            contracts.append(item.name)
            for json_file in item.glob("*.json"):
                models.add(json_file.stem)
    
    return contracts, sorted(models)


def load_aggregated(base_path: Path, contract: str, model: str):
    """Load aggregated JSON for contract/model."""
    path = base_path / contract / f"{model}.json"
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


def apply_header_style(ws, row, cols):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=50):
    """Auto-adjust column widths."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row + 1, 100)):  # Sample first 100 rows
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def generate_master_workbook(base_path: Path, contracts: list, models: list, output_path: Path):
    """Generate master workbook with all contract/model results."""
    wb = Workbook()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # === SUMMARY RANKINGS ===
    ws = wb.active
    ws.title = "Summary Rankings"
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "LEAH MODEL EVALUATION - TEST RESULTS"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:P2')
    ws['A2'] = f"Generated: {timestamp}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ["Rank", "Contract", "Model", "Raw Score", "Penalty", "Final Score", 
               "Det Pts", "Qual Pts", "T1 Gate", "T1%", "T2%", "T3%", "Y", "P", "N", "NMI"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, len(headers))
    
    # Collect all results
    results = []
    for contract in contracts:
        for model in models:
            agg = load_aggregated(base_path, contract, model)
            if agg and 'summary' in agg:
                s = agg['summary']
                det_by_tier = s.get('detection_by_tier', {})
                det_counts = s.get('detection_counts', {})
                
                results.append({
                    'contract': contract.replace('_', ' ').title(),
                    'model': model.replace('_', ' ').title(),
                    'raw_score': s.get('total_points', 0),
                    'penalty': s.get('penalty_pct', 0),
                    'final_score': s.get('final_score', s.get('total_points', 0)),
                    'det_pts': s.get('total_detection_points', 0),
                    'qual_pts': s.get('total_quality_points', 0),
                    't1_pass': s.get('t1_all_detected', False),
                    't1_pct': det_by_tier.get('T1', {}).get('pct', 0),
                    't2_pct': det_by_tier.get('T2', {}).get('pct', 0),
                    't3_pct': det_by_tier.get('T3', {}).get('pct', 0),
                    'y': det_counts.get('Y', 0),
                    'partial': det_counts.get('P', det_counts.get('△', 0)),
                    'n': det_counts.get('N', 0),
                    'nmi': det_counts.get('NMI', 0),
                    'contract_raw': contract,
                    'model_raw': model
                })
    
    # Sort by final score
    results.sort(key=lambda x: x['final_score'], reverse=True)
    
    for i, r in enumerate(results, 1):
        row = i + 4
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r['contract'])
        ws.cell(row=row, column=3, value=r['model'])
        ws.cell(row=row, column=4, value=round(r['raw_score'], 2))
        ws.cell(row=row, column=5, value=f"{r['penalty']}%" if r['penalty'] else "-")
        ws.cell(row=row, column=6, value=round(r['final_score'], 2))
        ws.cell(row=row, column=7, value=round(r['det_pts'], 1))
        ws.cell(row=row, column=8, value=round(r['qual_pts'], 1))
        
        t1_cell = ws.cell(row=row, column=9, value="PASS" if r['t1_pass'] else "FAIL")
        t1_cell.fill = PASS_FILL if r['t1_pass'] else FAIL_FILL
        
        ws.cell(row=row, column=10, value=f"{r['t1_pct']:.0f}%")
        ws.cell(row=row, column=11, value=f"{r['t2_pct']:.0f}%")
        ws.cell(row=row, column=12, value=f"{r['t3_pct']:.0f}%")
        ws.cell(row=row, column=13, value=r['y'])
        ws.cell(row=row, column=14, value=r['partial'])
        ws.cell(row=row, column=15, value=r['n'])
        ws.cell(row=row, column=16, value=r['nmi'])
        
        for col in range(1, 17):
            ws.cell(row=row, column=col).border = THIN_BORDER
    
    auto_width(ws)
    
    # === MODEL AGGREGATES ===
    ws2 = wb.create_sheet("Model Aggregates")
    ws2['A1'] = "Model Performance Across All Contracts"
    ws2['A1'].font = TITLE_FONT
    
    headers = ["Model", "Contracts", "Total Final", "Avg Final", "Total Raw", 
               "Penalties Applied", "T1 Pass Rate", "Avg T1%", "Avg T2%", "Avg T3%", "Total Y", "Total N"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=header)
    apply_header_style(ws2, 3, len(headers))
    
    # Aggregate by model
    model_agg = {}
    for r in results:
        m = r['model']
        if m not in model_agg:
            model_agg[m] = {'scores': [], 'raw': [], 't1_pass': 0, 't1_pct': [], 't2_pct': [], 't3_pct': [], 
                           'y': 0, 'n': 0, 'penalties': 0, 'count': 0}
        model_agg[m]['scores'].append(r['final_score'])
        model_agg[m]['raw'].append(r['raw_score'])
        model_agg[m]['t1_pass'] += 1 if r['t1_pass'] else 0
        model_agg[m]['t1_pct'].append(r['t1_pct'])
        model_agg[m]['t2_pct'].append(r['t2_pct'])
        model_agg[m]['t3_pct'].append(r['t3_pct'])
        model_agg[m]['y'] += r['y']
        model_agg[m]['n'] += r['n']
        model_agg[m]['penalties'] += 1 if r['penalty'] else 0
        model_agg[m]['count'] += 1
    
    sorted_models = sorted(model_agg.items(), key=lambda x: sum(x[1]['scores']), reverse=True)
    
    for i, (model, ma) in enumerate(sorted_models, 1):
        row = i + 3
        ws2.cell(row=row, column=1, value=model)
        ws2.cell(row=row, column=2, value=ma['count'])
        ws2.cell(row=row, column=3, value=round(sum(ma['scores']), 2))
        ws2.cell(row=row, column=4, value=round(sum(ma['scores']) / ma['count'], 2))
        ws2.cell(row=row, column=5, value=round(sum(ma['raw']), 2))
        ws2.cell(row=row, column=6, value=ma['penalties'] if ma['penalties'] else "")
        ws2.cell(row=row, column=7, value=f"{ma['t1_pass'] / ma['count'] * 100:.0f}%")
        ws2.cell(row=row, column=8, value=f"{sum(ma['t1_pct']) / len(ma['t1_pct']):.0f}%")
        ws2.cell(row=row, column=9, value=f"{sum(ma['t2_pct']) / len(ma['t2_pct']):.0f}%")
        ws2.cell(row=row, column=10, value=f"{sum(ma['t3_pct']) / len(ma['t3_pct']):.0f}%")
        ws2.cell(row=row, column=11, value=ma['y'])
        ws2.cell(row=row, column=12, value=ma['n'])
        
        for col in range(1, 13):
            ws2.cell(row=row, column=col).border = THIN_BORDER
    
    auto_width(ws2)
    
    # === CONTRACT COMPARISON ===
    ws3 = wb.create_sheet("Contract Comparison")
    ws3['A1'] = "Contract Difficulty Analysis"
    ws3['A1'].font = TITLE_FONT
    
    headers = ["Contract", "Best Model", "Best Score", "Worst Model", "Worst Score", "Spread", "Avg Score", "All T1 Pass"]
    for col, header in enumerate(headers, 1):
        ws3.cell(row=3, column=col, value=header)
    apply_header_style(ws3, 3, len(headers))
    
    # Aggregate by contract
    contract_results = {}
    for r in results:
        c = r['contract']
        if c not in contract_results:
            contract_results[c] = []
        contract_results[c].append(r)
    
    row = 4
    for contract, cr in sorted(contract_results.items(), key=lambda x: max(r['final_score'] for r in x[1]), reverse=True):
        cr.sort(key=lambda x: x['final_score'], reverse=True)
        best = cr[0]
        worst = cr[-1]
        avg = sum(r['final_score'] for r in cr) / len(cr)
        all_pass = all(r['t1_pass'] for r in cr)
        
        ws3.cell(row=row, column=1, value=contract)
        ws3.cell(row=row, column=2, value=best['model'])
        ws3.cell(row=row, column=3, value=round(best['final_score'], 2))
        ws3.cell(row=row, column=4, value=worst['model'])
        ws3.cell(row=row, column=5, value=round(worst['final_score'], 2))
        ws3.cell(row=row, column=6, value=round(best['final_score'] - worst['final_score'], 2))
        ws3.cell(row=row, column=7, value=round(avg, 2))
        
        pass_cell = ws3.cell(row=row, column=8, value="Yes" if all_pass else "No")
        pass_cell.fill = PASS_FILL if all_pass else FAIL_FILL
        
        for col in range(1, 9):
            ws3.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    auto_width(ws3)
    
    # === PENALTY IMPACT ===
    ws4 = wb.create_sheet("Penalty Impact")
    ws4['A1'] = "Extra Step Penalty Analysis"
    ws4['A1'].font = TITLE_FONT
    
    headers = ["Contract", "Model", "Raw Score", "Penalty", "Final Score", "Points Lost"]
    for col, header in enumerate(headers, 1):
        ws4.cell(row=3, column=col, value=header)
    apply_header_style(ws4, 3, len(headers))
    
    penalized = [r for r in results if r['penalty']]
    row = 4
    for r in sorted(penalized, key=lambda x: x['raw_score'] - x['final_score'], reverse=True):
        ws4.cell(row=row, column=1, value=r['contract'])
        ws4.cell(row=row, column=2, value=r['model'])
        ws4.cell(row=row, column=3, value=round(r['raw_score'], 2))
        ws4.cell(row=row, column=4, value=f"{r['penalty']}%")
        ws4.cell(row=row, column=5, value=round(r['final_score'], 2))
        ws4.cell(row=row, column=6, value=round(r['raw_score'] - r['final_score'], 2))
        
        for col in range(1, 7):
            ws4.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    auto_width(ws4)
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return results


def generate_contract_workbook(base_path: Path, contract: str, models: list, output_path: Path):
    """Generate detailed workbook for a single contract."""
    models_data = {}
    for model in models:
        agg = load_aggregated(base_path, contract, model)
        if agg:
            models_data[model] = agg
    
    if not models_data:
        return None
    
    wb = Workbook()
    
    # === SUMMARY ===
    ws = wb.active
    ws.title = "Summary"
    
    ws.merge_cells('A1:M1')
    ws['A1'] = f"{contract.upper()} - MODEL COMPARISON"
    ws['A1'].font = TITLE_FONT
    
    headers = ["Rank", "Model", "Total", "Det Pts", "Qual Pts", "T1", "T1%", "T2%", "T3%", "Y", "P", "N", "NMI"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, len(headers))
    
    results = []
    for model, agg in models_data.items():
        s = agg.get('summary', {})
        det_by_tier = s.get('detection_by_tier', {})
        det_counts = s.get('detection_counts', {})
        
        results.append({
            'model': model,
            'total': s.get('total_points', 0),
            'det': s.get('total_detection_points', 0),
            'qual': s.get('total_quality_points', 0),
            't1_pass': s.get('t1_all_detected', False),
            't1_pct': det_by_tier.get('T1', {}).get('pct', 0),
            't2_pct': det_by_tier.get('T2', {}).get('pct', 0),
            't3_pct': det_by_tier.get('T3', {}).get('pct', 0),
            'y': det_counts.get('Y', 0),
            'partial': det_counts.get('P', det_counts.get('△', 0)),
            'n': det_counts.get('N', 0),
            'nmi': det_counts.get('NMI', 0)
        })
    
    results.sort(key=lambda x: x['total'], reverse=True)
    
    for i, r in enumerate(results, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r['model'])
        ws.cell(row=row, column=3, value=round(r['total'], 2))
        ws.cell(row=row, column=4, value=round(r['det'], 1))
        ws.cell(row=row, column=5, value=round(r['qual'], 1))
        
        t1_cell = ws.cell(row=row, column=6, value="PASS" if r['t1_pass'] else "FAIL")
        t1_cell.fill = PASS_FILL if r['t1_pass'] else FAIL_FILL
        
        ws.cell(row=row, column=7, value=f"{r['t1_pct']:.0f}%")
        ws.cell(row=row, column=8, value=f"{r['t2_pct']:.0f}%")
        ws.cell(row=row, column=9, value=f"{r['t3_pct']:.0f}%")
        ws.cell(row=row, column=10, value=r['y'])
        ws.cell(row=row, column=11, value=r['partial'])
        ws.cell(row=row, column=12, value=r['n'])
        ws.cell(row=row, column=13, value=r['nmi'])
        
        for col in range(1, 14):
            ws.cell(row=row, column=col).border = THIN_BORDER
    
    auto_width(ws)
    
    # === GT MATRIX ===
    ws2 = wb.create_sheet("GT Matrix")
    
    first_model = list(models_data.values())[0]
    gt_items = first_model.get('gt_evaluations', [])
    
    headers = ["GT ID", "Clause", "Tier", "Issue"] + list(models_data.keys())
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
    apply_header_style(ws2, 1, len(headers))
    
    for i, gt in enumerate(gt_items):
        row = i + 2
        ws2.cell(row=row, column=1, value=gt.get('gt_id', ''))
        ws2.cell(row=row, column=2, value=gt.get('clause', ''))
        ws2.cell(row=row, column=3, value=gt.get('tier', ''))
        ws2.cell(row=row, column=4, value=gt.get('issue', '')[:50])
        
        tier_fill = TIER_FILLS.get(gt.get('tier'))
        if tier_fill:
            for c in range(1, 5):
                ws2.cell(row=row, column=c).fill = tier_fill
        
        for j, (model, agg) in enumerate(models_data.items(), 5):
            gt_evals = agg.get('gt_evaluations', [])
            m_gt = gt_evals[i] if i < len(gt_evals) else {}
            det = m_gt.get('detection', 'NMI')
            pts = m_gt.get('total_points', 0)
            
            cell = ws2.cell(row=row, column=j, value=f"{det} ({pts:.1f})")
            det_fill = DETECTION_FILLS.get(det)
            if det_fill:
                cell.fill = det_fill
            cell.alignment = Alignment(horizontal='center')
        
        for c in range(1, 5 + len(models_data)):
            ws2.cell(row=row, column=c).border = THIN_BORDER
    
    auto_width(ws2, max_width=40)
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Generate evaluation workbooks from aggregated data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python generate_workbooks.py --base-path ./aggregated
    python generate_workbooks.py --base-path ./aggregated --output-dir ./workbooks
    python generate_workbooks.py --base-path ./aggregated --contract consulting
    python generate_workbooks.py --base-path ./aggregated --master-only
        """
    )
    parser.add_argument('--base-path', '-b', type=Path, required=True,
                        help='Path to aggregated evaluation data')
    parser.add_argument('--output-dir', '-o', type=Path,
                        help='Output directory (default: {base-path}/../workbooks)')
    parser.add_argument('--contract', '-c',
                        help='Generate workbook for specific contract only')
    parser.add_argument('--master-only', action='store_true',
                        help='Only generate master workbook')
    parser.add_argument('--skip-master', action='store_true',
                        help='Skip master workbook, only generate per-contract')
    args = parser.parse_args()
    
    if not args.base_path.exists():
        print(f"ERROR: Base path not found: {args.base_path}")
        exit(1)
    
    output_dir = args.output_dir or (args.base_path.parent / 'workbooks')
    
    contracts, models = discover_structure(args.base_path)
    if not contracts:
        print("ERROR: No aggregated results found")
        exit(1)
    
    print(f"Base path: {args.base_path}")
    print(f"Output dir: {output_dir}")
    print(f"Contracts: {len(contracts)}")
    print(f"Models: {len(models)}")
    print()
    
    # Filter to specific contract if requested
    if args.contract:
        if args.contract not in contracts:
            print(f"ERROR: Contract '{args.contract}' not found")
            print(f"Available: {', '.join(contracts)}")
            exit(1)
        contracts = [args.contract]
    
    # Master workbook
    if not args.skip_master:
        print("Generating MASTER_EVALUATION_WORKBOOK.xlsx...", end=" ")
        master_path = output_dir / "MASTER_EVALUATION_WORKBOOK.xlsx"
        results = generate_master_workbook(args.base_path, contracts, models, master_path)
        print("✓")
        
        if args.master_only:
            print(f"\nOutput: {master_path}")
            return
    
    # Per-contract workbooks
    for contract in contracts:
        filename = f"{contract.upper()}_Evaluations.xlsx"
        print(f"Generating {filename}...", end=" ")
        contract_path = output_dir / filename
        result = generate_contract_workbook(args.base_path, contract, models, contract_path)
        print("✓" if result else "✗ (no data)")
    
    print(f"\n{'='*60}")
    print("WORKBOOKS GENERATED")
    print(f"{'='*60}")
    print(f"Output: {output_dir}")


if __name__ == "__main__":
    main()
